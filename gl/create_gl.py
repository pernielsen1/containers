
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.styles import Alignment


class accounting():
    COMPANY_SHEET = "firma"
    def __init__(self, config_file="config.json"):
        with open(config_file, 'r') as f:
            self.config = json.load(f)


        self.file_name = self.config['in_dir'] + "/" + "gl.xlsx"
        self.reports = pd.read_excel(self.file_name, sheet_name="rapporter")
        self.reports = self.reports.rename(columns={'beskrivning':'rapport_beskrivning'})
        self.company_excel = pd.read_excel(self.file_name, sheet_name=self.COMPANY_SHEET)           
        self.company_row= self.company_excel.iloc[0]

        self.accounts =  pd.read_excel(self.file_name, sheet_name="konti")
        self.accounts =  self.accounts[['konto', 'beskrivning', 'konto_typ']]  # don't take the formula fields 
        self.accounts_raw = self.accounts.copy(deep = True)
        self.accounts =  self.accounts.rename(columns={'beskrivning':'konto_beskrivning'})

        self.account_types =  pd.read_excel(self.file_name, sheet_name="konto_typer")
        self.account_types = self.account_types[['konto_typ', 'beskrivning', 'klass']] # don't take the formula fields
        self.account_types_raw  = self.account_types.copy(deep = True)
        self.account_types = self.account_types.rename(columns={'beskrivning':'konto_typ_beskrivning'})
        # add account types to accounts 
        self.accounts = self.accounts.merge(self.account_types, how='inner', 
                    left_on='konto_typ', right_on='konto_typ', indicator=False)
        self.classes = pd.read_excel(self.file_name, sheet_name="klasser")
        self.classes = self.classes[['klass', 'beskrivning', 'rapport_id']]
        self.classes_raw = self.classes.copy(deep=True)
        self.classes = self.classes.rename(columns={'beskrivning':'klass_beskrivning'})
        # add classes to accounts       
        self.accounts = self.accounts.merge(self.classes, how='inner', 
                    left_on='klass', right_on='klass', indicator=False)
        # add report id to accounts      
        self.accounts = self.accounts.merge(self.reports, how='inner', 
                    left_on='rapport_id', right_on='rapport_id', indicator=False)
        self.verifications = pd.read_excel(self.file_name, sheet_name='verifikationer')       
        self.verifications = self.verifications[['ver_nr', 'dato', 'beskrivning']]
        # read the first posting and calculate end of year
        year = str(self.verifications.iloc[0]['dato'])[0:4] 
        self.end_of_year = str(year) + '-12-31'
    
        self.postings_all =  pd.read_excel(self.file_name, sheet_name="posteringer")
        self.postings_all = self.postings_all[['ver_nr', 'lin', 'konto', 'belopp']] 
        # output files
        self.output_file = self.config['out_dir'] + "/" + "gl_output_" + str(year) + ".xlsx"
        self.wb = Workbook()
        self.wb_next_year = Workbook()
        self.next_year_file = self.config['out_dir'] + "/" + "gl_next_year.xlsx"
    
    def build_balances(self):
        postings_open = self.postings_all.query('ver_nr == 0') # opening balances
        postings = self.postings_all.query('ver_nr != 0')  # this year
        # create opening balance per account    
        account_open_balance = postings_open.groupby(by=['konto']).agg(
            opening_balance = ('belopp', 'sum'),
        )
        # movement opening balance per account    
        account_movements = postings.groupby(by=['konto']).agg(
            movement = ('belopp', 'sum'),
        )
        # outer join not all accounts will have open balance or movements
        balances = account_open_balance.merge(account_movements, how='outer', 
                                                   left_on='konto', right_on='konto', 
                                                   indicator=True)
        balances= balances.fillna({'opening_balance':0,'movement':0})
        balances['closing_balance'] = balances['opening_balance'] + balances['movement']
        # enrich with descriptions 
        balances = balances[['opening_balance', 'movement', 'closing_balance']]
        balances = self.accounts.merge(balances, how="left", 
                                            left_on='konto', right_on='konto', 
                                            indicator = True) 
        balances = balances.fillna({'opening_balance':0,'movement':0, 'closing_balance':0})      
        balances = balances.sort_values(['klass', 'konto_typ', 'konto'])
        balances = balances[['rapport_id', 'rapport_beskrivning', 'sheet_name',
                                        'klass', 'klass_beskrivning', 
                                       'konto_typ', 'konto_typ_beskrivning',
                                       'konto', 'konto_beskrivning',
                                       'opening_balance', 'movement', 'closing_balance']]
        return balances
    
    def run(self):
    
        balances  = self.build_balances()
        self.transfer_last_year(balances)
        self.book_result(balances)
        balances  = self.build_balances()  # after the result has been booked

        postings = self.postings_all.query('ver_nr != 0')  # this year
 
        # add data like text and date from verifications
        postings = postings.merge(self.verifications,  how="left", 
                                            left_on='ver_nr', right_on='ver_nr', 
                                            indicator = False) 
        postings = postings.merge(self.accounts, how="left",
                                            left_on='konto', right_on='konto', 
                                            indicator = False) 
                                   
        self.balance_report_to_excel(balances, 'BS')
        self.balance_report_to_excel(balances, 'RS')
        self.balance_report_to_excel(balances, 'ALL')  # balance list
        self.gl_report_to_excel(balances, postings, 'GL')  # GL
        
        self.verifikation_to_excel(postings)
        self.init_next_year(balances)
        self.wb.save(filename = self.output_file)
        self.wb_next_year.save(filename = self.next_year_file)
        print("all done - workbook ready")

    def set_numeric_cell(self, ws, row, column, value):
        ws.cell(row=row, column=column).value = value
        ws.cell(row=row, column=column).number_format = "0.00"

    def apply_format_to_column(self, ws, col, format, empty_col = 0):
        r = 1
        if empty_col == 0:
            empty_col = col
        while ws.cell(row=r, column=empty_col).value != None:
            ws.cell(row=r, column=col).number_format = format 
            r += 1
   
    def init_next_year(self, balances):
        next_year_start_date = str(int(self.end_of_year[0:4]) + 1) + '-01-01' 
        next_year_verifications = self.verifications.query('ver_nr == 0').copy(deep=True)
        next_year_verifications['dato'] = next_year_start_date
        ws = self.df_to_excel(self.wb_next_year, next_year_verifications, 'verifikationer')
        ws.cell(row=1, column=4).value = 'sum'
    #    ws.cell(row=2, column=4).value = '=_xlfn.SUMIFS(posteringer!D:D,posteringer!A:A,"="&A2)'
        ws.cell(row=2, column=4).value = '=_xlfn.ROUND(_xlfn.SUMIFS(posteringer!D:D,posteringer!A:A,"="&A2),2)'

        ws = self.df_to_excel(self.wb_next_year, self.accounts_raw, 'konti')
        ws.cell(row=1, column=4).value = 'konto_typ_beskrivning'
        ws.cell(row=1, column=5).value = 'klass'
        ws.cell(row=1, column=6).value = 'klass_beskrivning'
        ws.cell(row=1, column=7).value = 'ing_balans'
        ws.cell(row=1, column=8).value = 'perioden'
        ws.cell(row=1, column=9).value = 'utg_balans'
        for cur_row in range(len(self.accounts_raw)):
            r = cur_row + 2
            ws.cell(row=r, column = 4).value = f'=_xlfn.XLOOKUP(C{r},konto_typer!A:A,konto_typer!B:B)'
            ws.cell(row=r, column = 5).value = f'=_xlfn.XLOOKUP(C{r},konto_typer!A:A,konto_typer!C:C)'
            ws.cell(row=r, column = 6).value = f'=_xlfn.XLOOKUP(E{r},klasser!A:A,klasser!B:B)'
            str_formula = f'=SUMIFS(posteringer!$D:$D,posteringer!$A:$A,"=0",posteringer!$C:$C,"="&$A{r})'
            ws.cell(row=r, column = 7).value = str_formula
            str_formula = f'=SUMIFS(posteringer!$D:$D,posteringer!$A:$A,"<>0",posteringer!$C:$C,"="&$A{r})'
            ws.cell(row=r, column = 8).value = str_formula
            ws.cell(row=r, column = 9).value = f'=G{r} + H{r}' 
            
      
        ws = self.df_to_excel(self.wb_next_year, self.account_types_raw, 'konto_typer')
        ws.cell(row=1, column=4).value = 'klass_beskrivning'
        for cur_row in range(len(self.account_types_raw)):
            r = cur_row + 2
            ws.cell(row=r, column = 4).value = f'=_xlfn.XLOOKUP(C{r},klasser!A:A,klasser!B:B)'
        

        ws=self.df_to_excel(self.wb_next_year, self.classes_raw, 'klasser')
        ws.cell(row=1, column=4).value = 'report_description'
        for cur_row in range(len(self.classes_raw)):
            r = cur_row + 2
            ws.cell(row=r, column = 4).value = f'=_xlfn.XLOOKUP(C{r},rapporter!A:A,rapporter!B:B)'
        
        self.df_to_excel(self.wb_next_year, self.reports, 'rapporter')
        next_year_bal =  balances.query('(rapport_id == "BS") & (closing_balance != 0) ')
        next_year = next_year_bal[['konto', 'closing_balance']].copy(deep=True)
        next_year = next_year.rename(columns={'closing_balance':'belopp'})
        next_year['ver_nr'] = 0
        next_year['lin'] = next_year.index
        next_year = next_year[['ver_nr', 'lin', 'konto', 'belopp']]
        next_year_ws = self.df_to_excel(self.wb_next_year, next_year, 'posteringer')
        next_year_ws.cell(row=1, column=5).value = 'beskrivning'
        next_year_ws.cell(row=1, column=6).value = 'dato'
        next_year_ws.cell(row=1, column=7).value = 'konto_beskrivning'
        for cur_row in range(len(next_year)):
            r = cur_row + 2
            str_formula = f'=_xlfn.XLOOKUP(A{r},verifikationer!A:A,verifikationer!C:C)'
            next_year_ws.cell(row=r, column = 5).value = str_formula
            next_year_ws.cell(row=r, column = 6).value = f'=_xlfn.XLOOKUP(A{r},verifikationer!A:A,verifikationer!B:B)'
            next_year_ws.cell(row=r, column = 7).value = f'=_xlfn.XLOOKUP(C{r},konti!A:A,konti!B:B)'
        
    def book_result(self, balances):
        result_df = balances.query('rapport_id=="RS"').agg({'movement' : ['sum']})
        result = result_df.iloc[0]['movement']
        new_ver_nr  = self.add_verification(self.end_of_year, 'Result')
        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 1, 2099, result]
        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 1, 8999, -result]
        
    def transfer_last_year(self, balances):
        # find balance for last years result
        last_year_result_1 = balances.query('konto == 2098').iloc[0]['closing_balance']
        last_year_result_2 = balances.query('konto == 2099').iloc[0]['closing_balance']

        new_ver_nr  = self.add_verification(self.end_of_year, 'transfer')

        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 1, 2098, -last_year_result_1]
        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 2, 2091, +last_year_result_1]
        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 3, 2099, -last_year_result_2]
        self.postings_all.loc[len(self.postings_all)] = [new_ver_nr, 4, 2098, +last_year_result_2]
    
    def add_verification(self, date_str, text:str):
        new_ver_nr = 1 + self.verifications.agg({'ver_nr' : ['max']}).iloc[0]['ver_nr']
        ver_date = datetime.strptime(date_str, '%Y-%m-%d')
        self.verifications.loc[len(self.verifications)] = [new_ver_nr, ver_date, text]
        return new_ver_nr 

    def df_to_excel(self, wb, df, sheet_name):
        ws = self.add_or_get_ws(wb, sheet_name)
        col = 1
        for column in df.columns:
            ws.cell(row=1, column = col).value = column
            col += 1
        cur_row = 2
        for index, row in df.iterrows():
            col = 1
            for column in df.columns:
                ws.cell(row=cur_row, column=col).value = row[column]
                col += 1
            cur_row += 1
        return ws


        
    def write_line(self, ws, r, text, konto=None, opening=None, movement=None):
        if konto is not None:
            ws.cell(row=r, column=1).value = konto
        ws.cell(row=r, column=2).value= text
        if (opening is not None or movement is not None):
            self.set_numeric_cell(ws=ws, row=r, column=3, value=opening)
            self.set_numeric_cell(ws=ws, row=r, column=4, value=movement)
            self.set_numeric_cell(ws=ws, row=r, column=5, value=opening + movement)
        return r+1

    def write_dashes(self, ws, r, num_cols, dash_char):
        for col in range(num_cols): 
            f = "=REPT(" + '"' + dash_char + '"' + ',' + '10' + ')'
            ws.cell(row=r, column=col+1).value = f
        return r + 1


    def write_heading(self, ws, headings, report_name):
        ws.oddHeader.left.text = self.company_row['name']
        ws.oddHeader.center.text = report_name
        ws.oddHeader.right.text = "Sida &[Page]"
        ws.cell(row=1, column=1).value = self.company_row['name']
        ws.cell(row=1, column=3).value = report_name
        ws.cell(row=1, column=5).value = self.company_row['name']
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
        ws.cell(row=1, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=3).alignment = Alignment(horizontal='center')
    
        ws.freeze_panes = "A2"
        ws.print_title_rows='1:3'
        r=2
        col = 1
        for item in headings:
            ws.cell(row=r, column=col).value=item
            col += 1
        return self.write_dashes(ws, r+1, len(headings), '=')
    
    def add_or_get_ws(self, wb, sheet_name, clear=True):
        if sheet_name in wb.sheetnames: 
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        if clear:
            ws.delete_rows(1, ws.max_row) # clear the sheet.
        return ws 
    
    def gl_report_to_excel(self, balances, verifications, report_id):
        this_report = self.reports.query(f'(rapport_id=="{report_id}")').iloc[0]
        ws = self.add_or_get_ws(self.wb, this_report['sheet_name'])
        headings = ["konto", "Text", "Ing balans", "Perioden", "Utg Balans" ]
        cur_row = self.write_heading(ws,headings, this_report['rapport_beskrivning'])
        for index, row in balances.iterrows():
            ws.cell(row=cur_row, column=1).value = 'Ing Balance'
            ws.cell(row=cur_row, column=2).value = row['konto']
            ws.cell(row=cur_row, column=3).value = row['konto_beskrivning']
            ws.cell(row=cur_row, column=4).value = row['opening_balance']
            self.set_numeric_cell(ws=ws, row=cur_row, column=4, value=row['opening_balance'])

            cur_row += 1
            strqry = f"konto=={row['konto']}"
            postings = verifications.query(strqry)
            for ver_index, ver_row in postings.iterrows():
                ws.cell(row=cur_row, column=2).value = ver_row['beskrivning']
                self.set_numeric_cell(ws=ws, row=cur_row, column=3, value=ver_row['belopp'])
                cur_row += 1

            ws.cell(row=cur_row, column=1).value = 'Utg Balance'
            self.set_numeric_cell(ws=ws, row=cur_row, column=4, value=row['closing_balance'])
            
            cur_row += 1
            cur_row = self.write_dashes(ws, cur_row, len(headings),'-')
 
    def balance_report_to_excel(self, in_df, report_id):
        str_query = f'(rapport_id=="{report_id}")'
        this_report = self.reports.query(str_query).iloc[0]
        ws = self.add_or_get_ws(self.wb, this_report['sheet_name'])

        if (report_id == 'ALL'): 
            df = in_df.copy(deep=True)
        else:
            df = in_df.query(str_query).copy(deep=True)

        headings = ["konto", "Text", "Ing balans", "Perioden", "Utg Balans" ]
        r = self.write_heading(ws,headings, this_report['rapport_beskrivning'])

        prv_klass = 0     
        cur_klass_beskrivning = ''
        tot_klass_opening = 0
        tot_klass_movement = 0

        prv_konto_typ = 0
        cur_konto_typ_beskrivning = ''
        tot_konto_typ_opening = 0
        tot_konto_typ_movement = 0 
        
        for index, row in df.iterrows():
            # write summaries if break
            if (row['konto_typ'] != prv_konto_typ):
                tot_klass_opening += tot_konto_typ_opening
                tot_klass_movement += tot_konto_typ_movement
                if (prv_konto_typ != 0):
                    if report_id != 'ALL':
                        r = self.write_line(ws, r, "S: A " + cur_konto_typ_beskrivning, None, 
                                        tot_konto_typ_opening, tot_konto_typ_movement)
                    tot_konto_typ_opening = 0
                    tot_konto_typ_movement = 0

            if (row['klass'] != prv_klass):
                if (prv_klass != 0):  # write klass summary 
                    r = self.write_line(ws, r, "S: A " + cur_klass_beskrivning, None,
                                     tot_klass_opening, tot_klass_movement)
                    tot_klass_opening = 0
                    tot_klass_movement = 0
                    r += 1 # make a blank iine
            
            # write headings if we are on new i.e. a break
            if row['klass'] != prv_klass:
                prv_klass = row['klass']
                cur_klass_beskrivning = row['klass_beskrivning']
                if report_id != 'ALL':
                    r = self.write_line(ws, r, cur_klass_beskrivning)
               
            if row['konto_typ'] != prv_konto_typ:
                prv_konto_typ = row['konto_typ']
                cur_konto_typ_beskrivning = row['konto_typ_beskrivning']
                if report_id != 'ALL':
                    r = self.write_line(ws, r, cur_konto_typ_beskrivning)
    
            opening = row['opening_balance']
            movement = row['movement']
            if (opening != 0 or movement != 0):
                r=self.write_line(ws, r, row['konto_beskrivning'], row['konto'], row['opening_balance'], row['movement'])
                tot_konto_typ_opening += opening
                tot_konto_typ_movement += movement

        # write last summary 
        if (report_id != 'ALL'):
            r = self.write_line(ws, r, "S: A " + cur_konto_typ_beskrivning,  None,
                        tot_konto_typ_opening, tot_konto_typ_movement)
        
        r = self.write_line(ws, r, "S: A " + cur_klass_beskrivning, None,  
                       tot_klass_opening, tot_klass_movement)
    
    def verifikation_to_excel(self, df):
        report_id = 'VER'
        this_report = self.reports.query(f'(rapport_id=="{report_id}")').iloc[0]
        ws = self.add_or_get_ws(self.wb, 'verifikationslista', clear=True)
        headings = ["ver", "konto ks", "Text", "Debet", "Kredit" ]
        col = 0
        r = self.write_heading(ws,headings, this_report['rapport_beskrivning'])

        cur_ver_nr = 0
        for index, row in df.iterrows():
            if cur_ver_nr != row['ver_nr']:
                if cur_ver_nr != 0:
                    r = self.write_dashes(ws, r, len(headings), '-')
                cur_ver_nr = row['ver_nr']
                ws.cell(row=r, column=1).value = row['ver_nr']
                posting_date = row['dato']
                posting_date_str = posting_date.strftime('%y%m%d')
                ws.cell(row=r, column=2).value = posting_date_str  
                ws.cell(row=r, column=3).value = row['beskrivning']
                r += 1

            ws.cell(row=r, column=2).value = row['konto']
            ws.cell(row=r, column=3).value = row['konto_beskrivning']
            if  row['belopp'] < 0:
                self.set_numeric_cell(ws=ws, row=r, column=5, value=row['belopp'])
            else:
                self.set_numeric_cell(ws=ws, row=r, column=4, value=row['belopp'])
            r += 1

        r = self.write_dashes(ws, r, len(headings), '-')

                
if __name__ == '__main__':
    accounting_obj = accounting("config.json")
    accounting_obj.run()
    